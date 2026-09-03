#!/usr/bin/env python3
"""Convert an .xlsx file to CSV (one file per sheet).

Usage:
    python xlsx2csv.py <input.xlsx> [output_dir]

If output_dir is omitted, prints the first sheet to stdout. Otherwise writes
one <sheetname>.csv per sheet into output_dir (created if missing). Cell
values are flattened to strings; formulas are read as their cached values.
"""
import csv
import os
import sys

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


def cell_str(c):
    if c is None:
        return ""
    v = c.value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def sheet_to_rows(ws):
    rows = []
    for row in ws.iter_rows():
        rows.append([cell_str(c) for c in row])
    # Trim trailing empty rows/cols for a clean CSV.
    while rows and not any(rows[-1]):
        rows.pop()
    if rows:
        maxlen = max(len(r) for r in rows)
        for r in rows:
            del r[maxlen:]
    return rows


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    wb = load_workbook(src, data_only=True, read_only=True)
    if len(sys.argv) < 3:
        ws = wb.worksheets[0]
        w = csv.writer(sys.stdout, lineterminator="\n")
        w.writerows(sheet_to_rows(ws))
        return
    outdir = sys.argv[2]
    os.makedirs(outdir, exist_ok=True)
    for ws in wb.worksheets:
        name = ws.title.replace("/", "_").replace("\\", "_")
        path = os.path.join(outdir, f"{name}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerows(sheet_to_rows(ws))
        print(f"wrote {path}")


if __name__ == "__main__":
    main()
